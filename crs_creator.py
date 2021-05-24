"""

This script performs...

"""

import os

from openpyxl import load_workbook
import convertapi
from tqdm import tqdm

import config


# Функция для записи свойств документа в диапазон ячеек
def unmerge_write_merge(doc, cells: str, property_to_write):
    try:
        start_cell = cells.split(':')[0]
        doc.unmerge_cells(cells)
        doc[start_cell] = property_to_write
        doc.merge_cells(cells)
    except AttributeError:
        print('{} is not an excel sheet!'.format(doc))

# Функция для создания нового титульника паспорта
def new_tit (file_dict, file_name, path_trm):
    pasport = load_workbook('tit_template.xlsx')
    pasport_tit = pasport['Cover Page']
# Наименование документа (рус.)
    doc_name_ru = (str(file_dict[file_name][2]).upper()).replace('\n', ' ')
    unmerge_write_merge(pasport_tit, 'H15:O15', doc_name_ru)
# Наименование документа (англ.)
    doc_name_en = (str(file_dict[file_name][3]).upper()).replace('\n', ' ')
    unmerge_write_merge(pasport_tit, 'H16:O16', doc_name_en)
# Ревизия поставщика
    doc_rev = file_dict[file_name][4]
    unmerge_write_merge(pasport_tit, 'H23:O23', doc_rev)
# Номер документа
    doc_number = file_dict[file_name][5]
    unmerge_write_merge(pasport_tit, 'H17:O17', doc_number)
# Наименование документа (рус+англ) в колонтитуле
    ru_en = doc_name_ru + '\n' + doc_name_en
    unmerge_write_merge(pasport_tit, 'D41:M41', ru_en)
# Номер документа в колонтитуле + ревизия поставщика
    no_rev = 'Номер документа / Document Number:\n' + doc_number +\
    '\n' + 'Редакция / Revision: ' + doc_rev
    unmerge_write_merge(pasport_tit, 'N42:O42', no_rev)
    
    def fill_revision_field(issue_list):
        issue = file_dict[file_name][0] # Цель выпуска
# Список дат из словаря документов
        date_list = file_dict[file_name][6]
# Длина списка дат до даты отправки текущего трансмиттела включительно
        date_list_len = len(date_list)
        k = 0 # Начальное значение счётчика
        while k < date_list_len:
            row = 37 - k
            row2 = 47 + k
            pasport_tit['B'+str(row)] = issue_list[k] # Ревизия (титул.)
            pasport_tit['C'+str(row)] = issue_list[k] # Ревизия (титул.)
            pasport_tit['E'+str(row)] = issue 
            pasport_tit['D'+str(row)] = date_list[date_list_len-k]  # Дата выпуска
            
            cells = 'F' + str(row) + ':H' + str(row)
            unmerge_write_merge(pasport_tit, cells, 'ChA') # Подготовил
            
            cells = 'I' + str(row) + ':J' + str(row)
            unmerge_write_merge(pasport_tit, cells, 'DZ') # Проверил
            
            pasport_tit['K'+str(row)] = 'SC' # Утвердил
            pasport_tit['C'+str(row2)] = issue_list[k] # Ревизия (стр.2)
            
            cells = 'D' + str(row2) + ':E' + str(row2)
            unmerge_write_merge(pasport_tit, cells, issue_list[k])
            
            pasport_tit['F'+str(row2)] = 'ALL' # Параграф (стр.2)
            
            cells = 'G' + str(row2) + ':O' + str(row2)
            if 'IFR' in issue:
                rev_description = 'Выпущено для рассмотрения / Issued for Review'
            else:
                rev_description = 'Выпущено для использования / Issued for Use'
            unmerge_write_merge(pasport_tit, cells, rev_description) # Описание редакций (стр.2)
            
            k += 1 # Инкремент счётчика
    
    if doc_rev in config.ifu_list:
        fill_revision_field(config.ifu_list)
    else:
        fill_revision_field(config.ifr_list)
            
    pasport.save(path_trm + file_name + '.xlsx')
    convertapi.api_secret = 'tXofbUp7gutBrMGR'
    convertapi.convert('pdf', {
        'File': path_trm + file_name + '.xlsx'
    }, from_format = 'xlsx').save_files(path_trm)
    

for file_name in file_dict:
    template = load_workbook('template.xlsx')
    template_sheet = template[template.get_sheet_names()[0]]
    
    unmerge_write_merge(template_sheet, 'E7:F7', trm_name)
    trm_date = file_dict[file_name][6][-1]
    template_sheet['I7'] = trm_date
    issue = file_dict[file_name][0]
    template_sheet['A22'] = issue
    doc_number = file_dict[file_name][5]
    template_sheet['B22'] = doc_number
    doc_class = file_dict[file_name][1]
    template_sheet['C22'] = doc_class
    
    doc_name_ru = str(file_dict[file_name][2])
    doc_name_en = str(file_dict[file_name][3])
    ru_en = doc_name_ru + '\n' + doc_name_en
    unmerge_write_merge(template_sheet, 'D22:E22', ru_en)
    
    doc_rev = file_dict[file_name][4]
    template_sheet['F22'] = doc_rev
    
    crs_name = str(file_name)+'_CRS.xlsx'
    crs_path = os.path.join(path_trm, crs_name)
    template.save(crs_path)
#    new_tit (file_dict, file_name, path_trm)