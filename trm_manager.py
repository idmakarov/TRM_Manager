import datetime
import fnmatch
import os
import re
import shutil
import sys
import warnings

import fitz
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from PyPDF2 import PdfFileReader
from xlrd import open_workbook

import config
from database import DataBase
from document import Document
from print_progress_bar import PrintProgressBar
from transmittal import Transmittal


class TrmManager:
    """
    The class is purposed for managing transmittals.

    Attributes
    ----------
    db : Database
    print_dir : str
        Directory where to collect documents for printing.
    trm_dir : str
        Directory where to look for transmittals.
    vdr_dir : str
        Directory where to look for VDR.

    Methods
    -------
    create_crs(cur_trm: Transmittal)
        Creates comment review sheet for every existing document in the transmittal; optionally creates title page.
    create_trm_inventory(cur_trm: Transmittal, send_date: str)
        Creates transmittal inventory files.
    fill_vdr_fields(self, cur_trm: Transmittal)
        Fills required fields in VDR for each document in the transmittal.
    parse_received_trm(cur_trm: Transmittal)
        Parses info in documents corresponding to the current transmittal.
    parse_trm_docs(cur_trm: Transmittal, send_date: str)
        Parses data in documents corresponding to the current transmittal.
    prepare_docs_for_printing(cur_vdr: Document)
        Performs all preparations needed for printing documents from received transmittals.
    update_db(is_received=False)
        Updates database by adding new TRMs and VDRs.
    update_paths_and_files(update_docs=False)
        Updates paths to TRMs and updates file list corresponding to the TRM.
    """
    def __init__(self, trm_dir: str, vdr_dir: str, print_dir: str):
        warnings.filterwarnings('ignore', category=UserWarning)

        self.__cfg = config.get_config()
        self.__trm_dir = os.path.abspath(trm_dir)
        self.__vdr_dir = os.path.abspath(vdr_dir)
        self.__print_dir = print_dir

        print('Trying to load database...')
        prc_dir = self.__cfg[0]['process_dir']
        self.__db_path = os.path.join(prc_dir, 'db.pickle')
        self.db = DataBase(self.__db_path)

    def __parse_files(self, db: DataBase, item_type: str, is_received=False):
        """
        Parses files in folders using known path and adds documents and transmittals to the database.

        Parameters
        ----------
        db : DataBase
        item_type : {'trm', 'vdr'}
            Type of a target object to be parsed.
        is_received : bool, default=False
            Flag to indicate whether transmittals is received or not.

        Returns
        -------
        None
        """
        def clear_name(raw_name: str):
            if '.xls' in raw_name:
                item_name = raw_name.split('.xls')[0]
            elif '.XLS' in raw_name:
                item_name = raw_name.split('.XLS')[0]
            else:
                item_name = re.sub(r'(?<=TRM-\d{5}).*$', r'', raw_name)

            item_name = item_name.strip()
            return item_name

        mask_dict = self.__cfg[1]

        mask = ''
        directory = ''

        if item_type == 'vdr':
            directory = self.__vdr_dir
            mask = mask_dict['vdr_mask']
        elif item_type == 'trm':
            directory = self.__trm_dir

            if is_received:
                mask = mask_dict['received_trm_mask']
            else:
                mask = mask_dict['send_trm_mask']

        for item in os.listdir(directory):
            if fnmatch.fnmatch(item, mask):
                path = os.path.join(directory, item)
                name = clear_name(item)
                if item_type == 'trm':
                    if os.path.isdir(path):
                        obj = Transmittal(name, path)
                        db.add_item(obj)
                        print('{} was added to DB'.format(obj.name))
                elif item_type == 'vdr':
                    if os.path.isfile(path):
                        obj = Document(name, path)
                        db.add_item(obj)
                        print('{} was added to DB'.format(obj.name))
                else:
                    print("ERROR: method supports only 'vdr' and 'trm' item_type!", file=sys.stderr)
                    break

    def update_db(self, is_received=False):
        """
        Updates database by adding new TRMs and VDRs.

        Parameters
        ----------
        is_received : bool, default=False
            Flag to indicate whether transmittals is received or not.

        Returns
        -------
        None
        """
        print('DB is updating now...')
        self.db.clear_db()
        self.__parse_files(self.db, 'vdr')
        self.__parse_files(self.db, 'trm', is_received)
        self.db.save_db()
        print('DB was successfully updated')

    def update_paths_and_files(self, update_docs=False):
        """
        Updates paths to TRMs and updates file list corresponding to the TRM.

        Parameters
        ----------
        update_docs : bool, default=False

        Returns
        -------
        None
        """
        self.db.update_db(update_docs)
        s = 'and documents ' if not update_docs else ''
        print('Paths {}in DB was successfully updated'.format(s))

    # Функция для нахождения номера строки,
    # в которой находятся данные по конкретному документу
    def __get_vdr_ind(self, xlsheet: Worksheet, doc_name: str):
        """
        Finds a VDR row index which contains selected document data.

        Parameters
        ----------
        xlsheet : Worksheet
            Worksheet of VDR file.
        doc_name : str
            Name or number of target document.

        Returns
        -------
        i : int
            Row index where target document info locates.
        """
        if 'ER' in doc_name:
            s = re.sub(r'_.*$', '', doc_name)
            s = re.sub(r'-(?=[\w\d]{2}-[\d]{4})', r'.', s)
            return self.__get_vdr_ind(xlsheet, s)
        else:
            for i in range(13, xlsheet.max_row+1):
                cur_cell_value = xlsheet.cell(row=i, column=41).value
                if cur_cell_value is not None and cur_cell_value == doc_name:
                    return i
        print('ERROR: there is no {} in VDR!'.format(doc_name), file=sys.stderr)
        return None

    @staticmethod
    def __find_issued_cols(xlsheet: Worksheet, revision: str):
        """
        Finds VDR column with issue info.

        Parameters
        ----------
        xlsheet : Worksheet
            Worksheet of VDR file.
        revision : str
            Revision of target document.
            It consists of a letter and a digit or of two digits, for example: 'A1', '01'.

        Returns
        -------
        issue_cols : list
            List of columns index.
        """
        is_ifr = re.search(r'[A-Z]\d', revision)
        if is_ifr:
            filter_ = r'rev.[A-Z]\d'
        else:
            filter_ = r'rev.\d\d'

        issue_cols = []
        str1 = 'issue for'
        for i in range(62, xlsheet.max_column+1):
            cell_value = xlsheet.cell(row=9, column=i).value
            if str1 in str(cell_value) and bool(re.search(filter_, cell_value)):
                issue_cols.append(i+2)
        return issue_cols

    def __find_vdr(self, phase: str):
        """
        Finds required VDR in DB using phase provided.

        Parameters
        ----------
        phase : str

        Returns
        -------
        vdr_path : str
            Path to VDR file.
        """
        mask = self.__cfg[1]['vdr_mask']
        mask = mask.split('*')[0] + phase + '*'
        item_name_list = self.db.get_item_names()
        vdr = None
        for item_name in item_name_list:
            if fnmatch.fnmatch(item_name, mask):
                id_ = item_name_list.index(item_name)
                vdr = self.db.get_item(id_)
                print('VDR {} was found'.format(vdr.name))
                return vdr.path
        if vdr is None:
            print('ERROR: VDR was not found!', file=sys.stderr)
            return None

    @staticmethod
    def __preprocess_str(s):
        """
        Assigns type 'str' to the input string and cuts escape sequences and spaces from both start and end
        of the string.

        Parameters
        ----------
        s :
            Input string.

        Returns
        -------
        prep_s : str
            Preprocessed string.
        """
        if s is None:
            return None
        return str(s).strip()

    def parse_trm_docs(self, cur_trm: Transmittal, send_date: str):
        """
        Parses data in documents corresponding to the current transmittal.

        Parameters
        ----------
        cur_trm : Transmittal
            Current transmittal.
        send_date : str
            Date when the transmittal is to be sent.

        Returns
        -------
        None
        """
        def get_doc_info(xlsheet: Worksheet, xlsheet_data: Worksheet, doc_name: str):
            """
            Gets selected document info in VDR provided.

            Parameters
            ----------
            xlsheet : Worksheet
                VDR file worksheet.
            xlsheet_data : Worksheet
                VDR file worksheet (data only).
            doc_name : str
                Name of the target document.

            Returns
            -------
            info_list : list
                List with document information required.
            is_changed : bool
                Whether VDR has been changed or not.
            """
            vdr_ind = self.__get_vdr_ind(xlsheet_data, doc_name)
            if vdr_ind is None:
                return vdr_ind, None
            # Цель выпуска документа
            doc_issue = self.__preprocess_str(xlsheet_data.cell(row=vdr_ind, column=55).value)
            # Класс документа
            doc_class = self.__preprocess_str(xlsheet_data.cell(row=vdr_ind, column=22).value)
            # Наименование документа (рус.)
            doc_name_ru = self.__preprocess_str(xlsheet_data.cell(row=vdr_ind, column=50).value)
            # Наименование документа (англ.)
            doc_name_en = self.__preprocess_str(xlsheet_data.cell(row=vdr_ind, column=49).value)
            # Ревизия поставщика
            revision = self.__preprocess_str(xlsheet_data.cell(row=vdr_ind, column=54).value)
            if '-' in revision:
                revision = 'A1'

            # Номер документа
            doc_number = self.__preprocess_str(xlsheet_data.cell(row=vdr_ind, column=41).value)

            # Функция для формирования списка дат
            def get_date_list(issue_list: list, issue_cols: list):
                """
                Collects dates of sending in list.

                Parameters
                ----------
                issue_list : list
                    All revisions list.
                issue_cols : list
                    Current document revisions list.

                Returns
                -------
                date_list : list
                    List with dates of sending transmittals.
                is_changed : bool
                    Whether VDR has been changed or not.
                """
                date_list_ = []
                is_changed = False
                ind = issue_list.index(revision)
                for index, date_col in enumerate(issue_cols[:ind+1]):
                    if index == ind:
                        xlsheet.cell(row=vdr_ind, column=date_col).value = datetime.datetime.strptime(send_date,
                                                                                                      r'%d.%m.%Y')
                        xlsheet.cell(row=vdr_ind, column=date_col + 1).value = cur_trm.name
                        date_list_.append(send_date)
                        is_changed = True
                    else:
                        cur_cell = xlsheet_data.cell(row=vdr_ind, column=date_col)
                        if cur_cell.value is not None:
                            rev_date = cur_cell.value
                            try:
                                date_list_.append(rev_date.strftime('%d.%m.%Y'))
                            except AttributeError:
                                print('ERROR: cell {} value in VDR is not a date! \
                                Please, be sure that you have entered the date \
                                and change its format to "Date".'.format(cur_cell.coordinate), file=sys.stderr)

                if not date_list_ or len(date_list_) < ind + 1:
                    date_col = issue_cols[ind]
                    cur_date = datetime.datetime.now().date()
                    xlsheet.cell(row=vdr_ind, column=date_col).value = cur_date
                    xlsheet.cell(row=vdr_ind, column=date_col+1).value = cur_trm.name
                    date_list_.append(cur_date.strftime('%d.%m.%Y'))
                    is_changed = True

                return date_list_, is_changed

            [ifr_list, ifu_list] = self.__cfg[2].values()
            issue_cols = self.__find_issued_cols(xlsheet_data, revision)
            # Соберём список дат отправленных трансмиттелов до текущей ревизии.
            # Если номер ревизии не букво-цифровой, то искать в столбцах 00, 01,..
            if revision in ifu_list:
                date_list, is_changed = get_date_list(ifu_list, issue_cols)
                # Иначе искать в столбцах А1, В1,..
            else:
                date_list, is_changed = get_date_list(ifr_list, issue_cols)

            # Код дисциплины (Марка комплекта)
            doc_discipline_code = self.__preprocess_str(xlsheet_data.cell(row=vdr_ind, column=38).value)
            # Тип кода документа
            doc_type_code = self.__preprocess_str(xlsheet_data.cell(row=vdr_ind, column=40).value)

            info_list = [doc_issue, doc_class, doc_name_ru, doc_name_en, revision, doc_number, date_list,
                         doc_discipline_code, doc_type_code]

            return info_list, is_changed

        print('TRM docs info is parsing...')
        vdr_tmp = self.__find_vdr(cur_trm.phase)
        if vdr_tmp is None:
            print('ERROR: there is no VDR for phase {}!'.format(cur_trm.phase), file=sys.stderr)
            return
        # Загрузим данный VDR, при этом считываем только значения в ячейках
        wb = load_workbook(vdr_tmp)
        sheet = wb['VDR']
        wb_data = load_workbook(vdr_tmp, data_only=True)
        sheet_data = wb_data['VDR']

        total = len(cur_trm.documents)
        bar = PrintProgressBar(start=0, total=total, prefix='Progress:', suffix='Complete', length=50)
        check = 0
        for doc in cur_trm.documents:
            doc_info, is_changed = get_doc_info(sheet, sheet_data, doc)
            if is_changed:
                check += 1
            if doc_info is None:
                bar.print_progress_bar()
                continue
            cur_trm.documents[doc] = doc_info
            bar.print_progress_bar()

        # i = 0
        # if check:
        #     while True:
        #         try:
        #             wb.save(vdr_tmp)
        #             print('{} was changed and saved.'.format(os.path.split(vdr_tmp)[1]))
        #             break
        #         except PermissionError:
        #             if not i:
        #                 print('Please, close the file {} !'.format(os.path.split(vdr_tmp)[1]))
        #                 i += 1
        wb.close()
        wb_data.close()
        print('TRM docs parsing was successfully ended')

    # Функция для записи свойств документа в диапазон ячеек
    @staticmethod
    def __unmerge_write_merge(doc: Worksheet, cells: str, property_to_write):
        """
        Writes document properties in multiple cells and then merges them.

        Parameters
        ----------
        doc : Worksheet
            All revisions list.
        cells : str
            Current document revisions list.
        property_to_write : atomic data type

        Returns
        -------
        None
        """
        try:
            start_cell = cells.split(':')[0]
            doc.unmerge_cells(cells)
            doc[start_cell] = property_to_write
            doc.merge_cells(cells)
        except AttributeError:
            print('{} is not an excel sheet!'.format(doc))

    def create_crs(self, cur_trm: Transmittal):
        """
        Creates comment review sheet for every existing document in the transmittal; optionally creates title page.

        Parameters
        ----------
        cur_trm : Transmittal
            Current transmittal.

        Returns
        -------
        None
        """
        # Функция для создания нового титульника паспорта
        def new_tit(file_dict: dict, file_name: str, path_trm: str):
            """
            Creates passport new title page.

            Parameters
            ----------
            file_dict : dict
                Files dictionary with required information.
            file_name : str
                Selected file name.
            path_trm : str
                Path to the current transmittal.

            Returns
            -------
            None
            """
            print(f'Title sheets are being added to {file_name}...')
            tit_path = os.path.join(self.__cfg[0]['templates_dir'], 'tit_template.xlsx')
            pasport = load_workbook(tit_path)
            pasport_tit = pasport['Cover Page']
            # Наименование документа (рус.)
            doc_name_ru = (str(file_dict[file_name][2]).upper()).replace('\n', ' ')
            self.__unmerge_write_merge(pasport_tit, 'H15:O15', doc_name_ru)
            # Наименование документа (англ.)
            doc_name_en = (str(file_dict[file_name][3]).upper()).replace('\n', ' ')
            self.__unmerge_write_merge(pasport_tit, 'H16:O16', doc_name_en)
            # Ревизия поставщика
            doc_rev = file_dict[file_name][4]
            self.__unmerge_write_merge(pasport_tit, 'H23:O23', doc_rev)
            # Номер документа
            doc_number = file_dict[file_name][5]
            self.__unmerge_write_merge(pasport_tit, 'H17:O17', doc_number)
            # Наименование документа (рус+англ) в колонтитуле
            ru_en = doc_name_ru + '\n' + doc_name_en
            self.__unmerge_write_merge(pasport_tit, 'D41:M41', ru_en)
            # Номер документа в колонтитуле + ревизия поставщика
            no_rev = 'Номер документа / Document Number:\n' + doc_number +\
            '\n' + 'Редакция / Revision: ' + doc_rev
            self.__unmerge_write_merge(pasport_tit, 'N42:O42', no_rev)

            def fill_revision_field(issue_list: list):
                """
                Fills revision fields in the passport new title page.

                Parameters
                ----------
                issue_list : list
                    All revisions list.

                Returns
                -------
                None
                """
                issue = file_dict[file_name][0]  # Цель выпуска
                # Список дат из словаря документов
                date_list = file_dict[file_name][6]
                # Длина списка дат до даты отправки текущего трансмиттела включительно
                date_list_len = len(date_list)
                k = 0  # Начальное значение счётчика
                while k < date_list_len:
                    row = 37 - k
                    row2 = 47 + k
                    pasport_tit['B'+str(row)] = issue_list[k]  # Ревизия (титул.)
                    pasport_tit['C'+str(row)] = issue_list[k]  # Ревизия (титул.)
                    pasport_tit['E'+str(row)] = issue
                    pasport_tit['D'+str(row)] = date_list[k]  # Дата выпуска

                    cells = 'F' + str(row) + ':H' + str(row)
                    self.__unmerge_write_merge(pasport_tit, cells, 'ChA')  # Подготовил

                    cells = 'I' + str(row) + ':J' + str(row)
                    self.__unmerge_write_merge(pasport_tit, cells, 'DZ')  # Проверил

                    pasport_tit['K'+str(row)] = 'SC'  # Утвердил
                    pasport_tit['C'+str(row2)] = issue_list[k]  # Ревизия (стр.2)

                    cells = 'D' + str(row2) + ':E' + str(row2)
                    self.__unmerge_write_merge(pasport_tit, cells, issue_list[k])

                    pasport_tit['F'+str(row2)] = 'ALL'  # Параграф (стр.2)

                    cells = 'G' + str(row2) + ':O' + str(row2)
                    if 'IFR' in issue:
                        rev_description = 'Выпущено для рассмотрения / Issued for Review'
                    else:
                        rev_description = 'Выпущено для использования / Issued for Use'
                    self.__unmerge_write_merge(pasport_tit, cells, rev_description)  # Описание редакций (стр.2)

                    k += 1  # Инкремент счётчика

            [ifr_list, ifu_list] = self.__cfg[2].values()
            if doc_rev in ifu_list:
                fill_revision_field(ifu_list)
            else:
                fill_revision_field(ifr_list)

            file_path = os.path.join(path_trm, file_name)
            passport_name = file_name + '.xlsx'
            if os.path.isdir(file_path):
                passport_path = os.path.join(file_path, passport_name)
            else:
                passport_path = os.path.join(path_trm, passport_name)
            pasport.save(passport_path)

            # xlApp = client.Dispatch("Excel.Application")
            # books = xlApp.Workbooks.Open()
            # ws = books.Worksheets[0]
            # ws.Visible = 1
            # ws.ExportAsFixedFormat(0, 'C:\\excel\\trial.pdf')

            # convertapi.api_secret = 'tXofbUp7gutBrMGR'
            # convertapi.convert('pdf', {
            #     'File': passport_path
            # }, from_format='xlsx').save_files(os.path.split(passport_path)[0])
            print('Title sheets were successfully added')

        print('CRS files are creating...')
        trm_name = cur_trm.name
        trm_path = cur_trm.path
        template_path = os.path.join(self.__cfg[0]['templates_dir'], 'template.xlsx')

        file_dict = cur_trm.documents
        total = len(cur_trm.documents)
        bar = PrintProgressBar(start=0, total=total, prefix='Progress:', suffix='Complete', length=50)
        for doc in file_dict:
            if file_dict[doc]:
                # Загрузим образец CRS
                template = load_workbook(template_path)
                template_sheet = template[template.sheetnames[0]]
                # Заполним нужные поля
                self.__unmerge_write_merge(template_sheet, 'E7:F7', trm_name)
                trm_date = file_dict[doc][6][-1]
                template_sheet['I7'] = trm_date
                doc_issue = file_dict[doc][0]
                template_sheet['A22'] = doc_issue
                doc_number = file_dict[doc][5]
                template_sheet['B22'] = doc_number
                doc_class = file_dict[doc][1]
                template_sheet['C22'] = doc_class

                doc_name_ru = str(file_dict[doc][2])
                doc_name_en = str(file_dict[doc][3])
                ru_en = doc_name_ru + '\n' + doc_name_en
                self.__unmerge_write_merge(template_sheet, 'D22:E22', ru_en)

                doc_rev = file_dict[doc][4]
                template_sheet['F22'] = doc_rev
                # Сохраним файл CRS в корень TRM или в папку с документом, если
                # файлы хранятся в отдельных папках
                crs_name = str(doc)+'_CRS.xlsx'
                file_path = os.path.join(trm_path, doc)
                if os.path.isfile(file_path + '.pdf'):
                    crs_path = os.path.join(trm_path, crs_name)
                else:
                    crs_path = os.path.join(file_path, crs_name)
                template.save(crs_path)
                # Если необходимо добавить титульник к паспорту
                # code_type_list = ['JH', 'LB']
                # if file_dict[doc][-1] in code_type_list:
                #     new_tit(file_dict, doc, trm_path)
            else:
                print('ERROR: {} has no info!'.format(doc), file=sys.stderr)
            bar.print_progress_bar()
        print('CRS files creation was successfully completed')

    def __get_sheet_format_from_pdf(self, pdf_file: PdfFileReader, return_format_list=False):
        """
        Gets sheets format from selected pdf file.

        Parameters
        ----------
        pdf_file: PdfFileReader
        return_format_list: bool, default=False
            Whether to return list with formats of each page or not.

        Returns
        -------
        text : str
            If return_format_list=False returns string of pages format (e.g. 'A3, A4').
        format_list : list
            If return_format_list=True returns list of formats.
        """
        precision = 0.2
        format_dict = self.__cfg[3]

        page_count = pdf_file.getNumPages()
        pages_w_h = [(float(pdf_file.getPage(i).mediaBox.getWidth() / 72),
                      float(pdf_file.getPage(i).mediaBox.getHeight() / 72)) for i in range(page_count)]
        format_list = []
        for w_h in pages_w_h:
            page_format = ''
            for fmt in format_dict:
                for fmt_value in format_dict[fmt]:
                    if abs(fmt_value[0] - w_h[0]) <= precision and abs(fmt_value[1] - w_h[1]) <= precision:
                        page_format = fmt
                    elif abs(fmt_value[0] - w_h[1]) <= precision and abs(fmt_value[1] - w_h[0]) <= precision:
                        page_format = fmt
            if not page_format:
                page_format = 'Other'

            format_list.append(page_format)

        if return_format_list:
            return format_list

        format_set_list = list(set(format_list))
        format_set_list.sort()

        text = ', '.join(format_set_list)
        return text

    @staticmethod
    def __open_pdf(cur_trm: Transmittal, doc_name: str, bar=None):
        """
        Tries to open pdf-file using path to the transmittal, if failed tries to open it in subfolder.

        Parameters
        ----------
        cur_trm : Transmittal
            Current transmittal.
        doc_name : str
            Selected document name.
        bar : PrintProgressBar, default=None
            Progress bar instance.

        Returns
        -------
        pdf : PdfFileReader
        file_path : str
            Path to selected document.
        """
        file_path = os.path.join(cur_trm.path, doc_name + '.pdf')
        try:
            pdf = PdfFileReader(file_path, strict=False)
        except FileNotFoundError:
            file_path = os.path.join(cur_trm.path, doc_name, doc_name + '.pdf')
            try:
                pdf = PdfFileReader(file_path, strict=False)
            except FileNotFoundError:
                if bar:
                    bar.print_progress_bar()
                print('ERROR: {}.pdf was not found in {}'.format(doc_name, cur_trm.name), file=sys.stderr)
                return None, None
        if cur_trm.documents[doc_name] is None:
            if bar:
                bar.print_progress_bar()
            print('ERROR: {} has no info!'.format(doc_name), file=sys.stderr)
            return None, None
        return pdf, file_path

    def create_trm_inventory(self, cur_trm: Transmittal, send_date: str):
        """
        Creates transmittal inventory files.

        Parameters
        ----------
        cur_trm : Transmittal
            Current transmittal.
        send_date : str
            Sending date.

        Returns
        -------
        None
        """
        def get_date_from_pdf(filename: str):
            """
            Parses date from pdf-file.

            Parameters
            ----------
            filename : str
                Path to the target file.

            Returns
            -------
            date : str
                Date of last revision.
            """
            month_dict = {
                'jan': '01',
                'feb': '02',
                'mar': '03',
                'apr': '04',
                'may': '05',
                'jun': '06',
                'jul': '07',
                'aug': '08',
                'sep': '09',
                'oct': '10',
                'nov': '11',
                'dec': '12'
            }

            with fitz.open(filename) as f:
                try:
                    page1_text = f.loadPage(0).getText('text')
                except:
                    print('ERROR: {} is empty!'.format(os.path.split(filename)[-1]), file=sys.stderr)
                    return ''
                try:
                    page2_text = f.loadPage(1).getText('text')
                except:
                    page2_text = ''

            def get_temp_date_list(page: str):
                """
                Parses dates from pdf-file.

                Parameters
                ----------
                page : str
                    Page contents.

                Returns
                -------
                temp_list : list
                    Temporary list of parsed dates.
                """
                temp_list = []
                date_format_ru = r'\d\d\.\d\d\.\d{4}'
                date_format_en = r'\d\d-\w{3}-\d{4}'
                match_ru = re.findall(date_format_ru, page)
                match_en = re.findall(date_format_en, page)
                if match_ru:
                    for date_time in match_ru:
                        temp_list.append(datetime.datetime.strptime(date_time, '%d.%m.%Y'))
                elif match_en:
                    for date_time in match_en:
                        month = re.findall(r'\w{3}', date_time)[0]
                        sub = r'.' + month_dict[month.lower()] + r'.'
                        date_time_ru = re.sub(r'-\w{3}-', sub, date_time)
                        temp_list.append(datetime.datetime.strptime(date_time_ru, '%d.%m.%Y'))
                return temp_list

            date_ = ''
            temp_list = get_temp_date_list(page1_text)
            if temp_list:
                date_ = max(temp_list).strftime('%d.%m.%Y')
            else:
                temp_list = get_temp_date_list(page2_text)
                if temp_list:
                    date_ = max(temp_list).strftime('%d.%m.%Y')
                else:
                    print('WARNING: {} has no date in first two pages!'.format(os.path.split(filename)[-1]))
            return date_

        tpl_path = self.__cfg[0]['templates_dir']
        trm_template_path = os.path.join(tpl_path, 'TRM_file_template.xlsx')
        csv_template_path = os.path.join(tpl_path, 'CSV_template.xlsx')
        csv_db_template_path = os.path.join(tpl_path, 'CSV_DB.xlsx')

        wb2 = load_workbook(trm_template_path, data_only=True)
        sheet2 = wb2[wb2.sheetnames[0]]
        wb3 = load_workbook(csv_template_path, data_only=True)
        sheet3 = wb3['Document Load']
        wb4 = load_workbook(csv_db_template_path, data_only=True, read_only=True)
        sheet4 = wb4['Лист1']
        sheet5 = wb4['Лист2']
        print('Templates reading was successfully completed')
        # Считаем в списки номера документов и типы кодов документов
        doc_number_list = []
        doc_type_code_list = []
        for i in range(1, sheet4.max_row+1):
            doc_number_list.append(sheet4.cell(row=i, column=1).value)
        for i in range(1, sheet5.max_row+1):
            doc_type_code_list.append(sheet5.cell(row=i, column=1).value)

        # Запишем выбранную пользователем дату и название трансмиттела в заголовок файла описи
        send_date_str = datetime.datetime.strptime(send_date, r'%d.%m.%Y')
        self.__unmerge_write_merge(sheet2, 'I1:J1', send_date_str)
        self.__unmerge_write_merge(sheet2, 'C2:J2', cur_trm.name)

        # Введём списки и словари для некоторых полей, чтобы дальше вставить в ячейку нужное описание
        issue_dict = {'IFR': 'IFR - Выпущено для рассмотрения',
                      'IFU': 'IFU - Выпущено для использования'}
        doc_class_dict = {'1': '1 - Обязательна приемка Генподрядчиком',
                          '2': '2 -Запрашивается рассмотрение Генподрядчиком',
                          '3': '3 - Приемка или рассмотрение Генподрядчиком не запрашивается'}
        package_type = {'ATH': ['ATH - Automated Systems', 'ATH - Автоматизированные системы'],
                        'ASK': ['ASK - Automatic Power Supply Control Systems',
                                'ASK - Автоматизированные системы управления электроснабжением'],
                        'PS': ['PS - Fire Alarm', 'PS - Пожарная сигнализация']}
        # Начнём заполнение для каждого документа
        print('Template files are being filled...')
        count = 7
        total = len(cur_trm.documents)
        bar = PrintProgressBar(start=0, total=total, prefix='Progress:', suffix='Complete', length=50)
        for doc_name in cur_trm.documents:
            # Заполним файл описи документов трансмиттела
            count += 1
            # Имя файла
            sheet2.cell(row=count, column=18).value = doc_name + '.pdf'
            # Расширение файла
            sheet2.cell(row=count, column=17).value = 'pdf'

            pdf, file_path = self.__open_pdf(cur_trm, doc_name, bar)
            if not pdf:
                continue

            # Формат страниц файла
            sheet_format = self.__get_sheet_format_from_pdf(pdf)
            sheet2.cell(row=count, column=16).value = sheet_format
            # Считаем файл и найдём число страниц
            page_count = pdf.getNumPages()
            sheet2.cell(row=count, column=15).value = page_count

            # Тип кода документа
            doc_type_code = cur_trm.documents[doc_name][8]
            sheet2.cell(row=count, column=14).value = doc_type_code
            # Ревизия поставщика
            doc_rev = cur_trm.documents[doc_name][4]
            sheet2.cell(row=count, column=13).value = doc_rev
            # Класс документа
            doc_class = cur_trm.documents[doc_name][1]
            sheet2.cell(row=count, column=12).value = doc_class
            # Дата ревизии документа

            doc_rev_date = get_date_from_pdf(file_path)
            sheet2.cell(row=count, column=11).value = doc_rev_date
            # Цель выпуска документа
            doc_issue = cur_trm.documents[doc_name][0]
            sheet2.cell(row=count, column=10).value = doc_issue
            # Код дисциплины (Марка комплекта)
            doc_disc_code = cur_trm.documents[doc_name][7]
            sheet2.cell(row=count, column=9).value = doc_disc_code
            # Наименование документа (англ.)
            doc_en_name = cur_trm.documents[doc_name][3]
            sheet2.cell(row=count, column=7).value = doc_en_name
            # Наименование документа (рус.)
            doc_ru_name = cur_trm.documents[doc_name][2]
            sheet2.cell(row=count, column=6).value = doc_ru_name
            sheet2.cell(row=count, column=5).value = 'ER'
            # Номер документа
            doc_number = cur_trm.documents[doc_name][5]
            sheet2.cell(row=count, column=4).value = doc_number
            # Заказ на покупку
            sheet2.cell(row=count, column=1).value = 'P2AM-7-0001'

            # Заполним файл CSV (опись трансмиттела)
            # Найдём нативный файл, соответствующий данному документу
            check = 0
            for sfile in os.listdir(cur_trm.path):
                if doc_name in sfile and os.path.isfile(os.path.join(cur_trm.path, sfile)):
                    if r'.pdf' not in sfile and 'CRS' not in sfile:
                        sheet3.cell(row=count-6, column=51).value = sfile
                        sheet3.cell(row=count-6, column=50).value = 'Native Format'
                        check = 1
                        break
                    elif r'.pdf' in sfile:
                        sheet3.cell(row=count - 6, column=51).value = sfile
                        sheet3.cell(row=count - 6, column=50).value = 'Native Format'
                        check = 1
                        break
            # Если в папке трансмиттела документы размещены в отдельных папках
            if not check:
                subfolder = os.path.join(cur_trm.path, doc_name)
                for sfile in os.listdir(subfolder):
                    if doc_name in sfile and os.path.isfile(os.path.join(subfolder, sfile)) and \
                            '.pdf' not in sfile and not 'CRS' in sfile:
                        sheet3.cell(row=count-6, column=51).value = sfile
                        sheet3.cell(row=count-6, column=50).value = 'Native Format'
                        check = 1
                        break
            # Если в папке трансмиттела нет данного документа
            if not check:
                print('WARNING: there is no {} in {}!'.format(doc_name, cur_trm.name))

            sheet3.cell(row=count-6, column=49).value = doc_name
            sheet3.cell(row=count-6, column=48).value = '00. Holding Folder'
            sheet3.cell(row=count-6, column=47).value = 'Latest'
            sheet3.cell(row=count-6, column=45).value = cur_trm.name
            sheet3.cell(row=count-6, column=43).value = sheet_format
            sheet3.cell(row=count-6, column=42).value = 'ER'
            sheet3.cell(row=count-6, column=36).value = page_count
            sheet3.cell(row=count-6, column=35).value = page_count
            if doc_disc_code == 'PS':
                sheet3.cell(row=count-6, column=32).value = 'SAT - Security and Telecommunication'
                sheet3.cell(row=count-6, column=31).value = 'SAT - Связь и телекоммуникации'
            else:
                sheet3.cell(row=count-6, column=32).value = 'CSY - Control systems'
                sheet3.cell(row=count-6, column=31).value = 'CSY - Системы управления'

            if doc_class is not None:
                doc_class_desc = doc_class_dict[doc_class]
                sheet3.cell(row=count-6, column=30).value = doc_class_desc
            else:
                sheet3.cell(row=count-6, column=30).value = ''

            sub_disc_code = re.sub(r'\d', '', doc_disc_code)
            sheet3.cell(row=count-6, column=28).value=package_type[sub_disc_code][0]
            sheet3.cell(row=count-6, column=27).value=package_type[sub_disc_code][1]

            # Добавим описание в 19-26 столбцы из файла 'CSV_DB.xlsx'
            doc_number_slice = doc_number[:25]
            for i in range(2, 10):
                sheet3.cell(row=count-6, column=i+17).value = sheet4.cell(row=doc_number_list.index(doc_number_slice)+1,
                                                                          column=i).value

            sheet3.cell(row=count-6, column=18).value = 'P2 ~ Нелицензионные установки'
            sheet3.cell(row=count-6, column=17).value = '4.' + cur_trm.phase
            sheet3.cell(row=count-6, column=16).value = '4 - ГПЗ'
            sheet3.cell(row=count-6, column=15).value = 'P2AM-7-0001-01'
            sheet3.cell(row=count-6, column=14).value = 'GAZprom Automation (1)'
            sheet3.cell(row=count-6, column=13).value = '0055'
            sheet3.cell(row=count-6, column=12).value = 'CPECC'
            try:
                sheet3.cell(row=count-6, column=11).value = issue_dict[doc_issue]
            except KeyError:
                pass
            sheet3.cell(row=count-6, column=10).value = doc_rev
            sheet3.cell(row=count-6, column=9).value = doc_rev_date
            sheet3.cell(row=count-6, column=8).value = sheet5.cell(row=doc_type_code_list.index(doc_type_code)+1,
                                                                   column=3).value
            sheet3.cell(row=count-6, column=7).value = sheet5.cell(row=doc_type_code_list.index(doc_type_code)+1,
                                                                   column=2).value
            sheet3.cell(row=count-6, column=6).value = doc_type_code
            sheet3.cell(row=count-6, column=5).value = doc_en_name
            sheet3.cell(row=count-6, column=4).value = doc_ru_name
            sheet3.cell(row=count-6, column=2).value = re.sub(r'_.*$', '', doc_name)
            bar.print_progress_bar()

        inventory_path = os.path.join(cur_trm.path, cur_trm.name + '.xlsx')
        inventory_csv_path = os.path.join(cur_trm.path, cur_trm.name + '_CSV.xlsx')
        wb2.save(inventory_path)
        wb3.save(inventory_csv_path)
        print('For {} inventory files were successfully created'.format(cur_trm.name))

    def __get_received_trm_inventory_path(self, trm_name: str, path: str):
        """
        Looks for transmittal inventory file using transmittal name and path provided.

        Parameters
        ----------
        trm_name : str
            Name of the transmittal.
        path : str
            Path to the transmittal folder.

        Returns
        -------
        path : str
            Path to the transmittal inventory file.
        """
        mask = trm_name + '.xls*'
        file_check = 0
        for item in os.listdir(path):
            file_check += 1
            if fnmatch.fnmatch(item, mask):
                path = os.path.join(path, item)
                return path
        if not file_check:
            print('ERROR: there are no files and folders in {}'.format(trm_name), file=sys.stderr)
        else:
            path = os.path.join(path, trm_name)
            try:
                self.__get_received_trm_inventory_path(trm_name, path)
            except:
                print('ERROR: TRM inventory file was not found!', file=sys.stderr)

    @staticmethod
    def clarify_doc_name(trm_doc_names, doc_name_to_clarify: str):
        """
        Compares document name from TRM inventory file and file name in corresponding folder.
        If parts of the names before 'ER' match then return file name, else - document name from TRM inventory file.

        Parameters
        ----------
        trm_doc_names : list-like
            List of files name in transmittal folder.
        doc_name_to_clarify : str
            Document name from TRM inventory file to be clarified.
        """
        trm_doc_names = list(trm_doc_names)
        name_1_list = [doc_name.split('ER')[0] for doc_name in trm_doc_names]
        name_2 = doc_name_to_clarify.split('ER')[0]

        try:
            index = name_1_list.index(name_2)
            return trm_doc_names[index]
        except ValueError:
            return doc_name_to_clarify

    def parse_received_trm(self, cur_trm: Transmittal):
        """
        Parses info in documents corresponding to the current transmittal.

        Parameters
        ----------
        cur_trm : Transmittal
            Current transmittal.

        Returns
        -------
        None
        """
        print('Processing {}...'.format(cur_trm.name))

        inventory_path = self.__get_received_trm_inventory_path(cur_trm.name, cur_trm.path)
        if inventory_path is None:
            return

        wb_data = open_workbook(inventory_path, logfile=open(os.devnull, 'w'))
        sheet_data = wb_data.sheet_by_index(0)

        col_names = [str(sheet_data.cell_value(6, i)).lower() for i in range(sheet_data.ncols)]
        try:
            doc_number_col = col_names.index([c for c in col_names if 'project doc number' in c.lower()][0])
            doc_name_col = col_names.index([c for c in col_names if 'electronic filename' in c.lower()][0])
            doc_rev_col = col_names.index('rev')
            crs_code_col = col_names.index('comments')
        except ValueError:
            print('ERROR: cannot parse trm inventory file (unknown columns name)!', file=sys.stderr)
            return

        rng = range(7, sheet_data.nrows)
        for i in rng:
            if sheet_data.cell_value(rowx=i, colx=3):
                doc_number = self.__preprocess_str(sheet_data.cell_value(rowx=i, colx=doc_number_col))
                doc_name = self.__preprocess_str(sheet_data.cell_value(rowx=i, colx=doc_name_col)[:-4])
                doc_rev = self.__preprocess_str(sheet_data.cell_value(rowx=i, colx=doc_rev_col))
                crs_code = self.__preprocess_str(sheet_data.cell_value(rowx=i, colx=crs_code_col))
                phase = cur_trm.phase
                trm_date = sheet_data.cell_value(rowx=0, colx=8)
                prop_list = [doc_number, doc_rev, crs_code, phase, trm_date]

                doc_name = self.clarify_doc_name(cur_trm.documents.keys(), doc_name)
                cur_trm.documents[doc_name] = prop_list

        print('Received TRM was successfully processed')

    def fill_vdr_fields(self, cur_trm: Transmittal):
        """
        Fills required fields in VDR for each document in the transmittal.

        Parameters
        ----------
        cur_trm : Transmittal
            Current transmittal.

        Returns
        -------
        None
        """
        def fill_doc_info(xlsheet: Worksheet, xlsheet_data: Worksheet, docs: dict):
            """
            Fills required fields in VDR using documents dictionary provided.

            Parameters
            ----------
            xlsheet : Worksheet
                VDR worksheet.
            xlsheet_data : Worksheet
                VDR worksheet (data only).
            docs : dict
                Documents dictionary with required information.

            Returns
            -------
            None
            """
            [ifr_list, ifu_list] = self.__cfg[2].values()

            total = len(docs)
            bar = PrintProgressBar(start=0, total=total, prefix='Progress:', suffix='Complete', length=50)
            for doc in docs:
                doc_num = docs[doc][0]
                # Вычислим номер строки, в которой находится нужный документ
                vdr_ind = self.__get_vdr_ind(xlsheet_data, doc_num)
                if vdr_ind is None:
                    bar.print_progress_bar()
                    continue
                # Вычислим номер столбца, с которого начнём заполнять информацию из полученного трансмиттела
                doc_rev = docs[doc][1]
                issue_cols = self.__find_issued_cols(xlsheet_data, doc_rev)
                if doc_rev in ifr_list:
                    ind = ifr_list.index(doc_rev)
                    req_col = issue_cols[ind]
                else:
                    ind = ifu_list.index(doc_rev)
                    req_col = issue_cols[ind]

                if doc_rev == 'A1':
                    req_col = req_col + 28
                else:
                    req_col = req_col + 4

                # Дата получения трансмиттела с замечаниями
                xlsheet.cell(row=vdr_ind, column=req_col).value = docs[doc][4]
                # Номер трансмиттела
                xlsheet.cell(row=vdr_ind, column=req_col+1).value = cur_trm.name
                # Код замечания CRS
                xlsheet.cell(row=vdr_ind, column=req_col+2).value = docs[doc][2]
                bar.print_progress_bar()

        print('Adding required fields into VDR...')
        vdr_tmp = self.__find_vdr(cur_trm.phase)
        if vdr_tmp is None:
            print('ERROR: there is no VDR for phase {}!'.format(cur_trm.phase), file=sys.stderr)
            return 1
        wb = load_workbook(vdr_tmp)
        sheet = wb['VDR']
        wb_data = load_workbook(vdr_tmp, data_only=True)
        sheet_data = wb_data['VDR']

        fill_doc_info(sheet, sheet_data, cur_trm.documents)
        # file_name = self.__working_dir + r'\vdr_' + cur_trm.phase + '.xlsx'
        wb.save(vdr_tmp)
        wb_data.close()
        print('Required fields were successfully added into VDR')

    def __parse_all_docs_info_from_vdr(self, cur_vdr: Document):
        """
        Collects documents information from vdr and assign specific status to each document.

        Parameters
        ----------
        cur_vdr : Document
            Current VDR.

        Returns
        -------
        doc_dict : dict
            Documents dictionary with parsed information.
        """
        doc_dict = {}

        print('Loading {}...'.format(cur_vdr.name))
        vdr_tmp = cur_vdr.path
        wb_data = load_workbook(vdr_tmp, data_only=True)
        sheet_data = wb_data['VDR']

        print('{} data was successfully read!'.format(cur_vdr.name))

        issue_cols = self.__find_issued_cols(sheet_data, '00')
        req_cols = [col + 4 for col in issue_cols]

        print('VDR info is being parsed...')

        rng = range(13, sheet_data.max_row + 1)
        total = len(rng)
        bar = PrintProgressBar(start=0, total=total, prefix='Progress:', suffix='Complete', length=50)
        for i in rng:
            cur_cell_value = sheet_data.cell(row=i, column=41).value

            if cur_cell_value and '0055' in cur_cell_value:
                doc_number = self.__preprocess_str(cur_cell_value)

                trm_name = ''
                crs_code = ''
                for req_col in req_cols:
                    trm_cell_value = sheet_data.cell(row=i, column=req_col + 1).value
                    if trm_cell_value:
                        trm_name = self.__preprocess_str(trm_cell_value)
                        crs_code = self.__preprocess_str(sheet_data.cell(row=i, column=req_col + 2).value)
                    else:
                        break
                if trm_name:
                    if crs_code == '1':
                        status = 'Ок'
                    elif crs_code is not None:
                        status = 'CRS-код актуальной ревизии - ' + crs_code
                    else:
                        status = 'CRS-код актуальной ревизии отсутствует'
                else:
                    status = 'Документ не был получен в статусе "выпущен для использования"'

                doc_dict[doc_number] = [trm_name, crs_code, status]

            bar.print_progress_bar()

        print('VDR info parsing was successfully ended!')

        wb_data.close()
        return doc_dict

    def __parse_docs_in_received_trms(self, doc_dict: dict):
        """
        Gets information about formats and pages number from each document (PDF-file) in received transmittal.

        Parameters
        ----------
        doc_dict : dict
            Documents dictionary with parsed information.

        Returns
        -------
        total_size : int
            Total size of files for printing in bytes.
        """
        item_names_list = self.db.get_item_names()
        trm_names_list = [item_name for item_name in item_names_list if 'TRM' in item_name]

        print('Parsing documents info in received transmittals...')

        for trm_name in trm_names_list:
            trm_id = item_names_list.index(trm_name)
            trm = self.db.get_item(trm_id)
            self.parse_received_trm(trm)

        print('Getting pages info from documents...')
        total = len(doc_dict)
        bar = PrintProgressBar(start=0, total=total, prefix='Progress:', suffix='Complete', length=50)

        total_size = 0

        for doc_num in doc_dict:
            trm_name = doc_dict[doc_num][0]
            status = doc_dict[doc_num][2]

            if trm_name in trm_names_list:
                trm_id = item_names_list.index(trm_name)
                cur_trm = self.db.get_item(trm_id)

                try:
                    doc_index = [
                        cur_trm.documents[doc][0]
                        if cur_trm.documents[doc]
                        else ''
                        for doc in cur_trm.documents
                    ].index(doc_num)
                    doc_name = list(cur_trm.documents)[doc_index]
                except ValueError:
                    print('ERROR: {} was not found in {}'.format(doc_num, cur_trm.name), file=sys.stderr)
                    bar.print_progress_bar()
                    continue

                pdf, file_path = self.__open_pdf(cur_trm, doc_name)
                if not pdf:
                    bar.print_progress_bar()
                    continue

                file_size = os.path.getsize(file_path)
                if status == 'Ок':
                    total_size += file_size

                format_list = self.__get_sheet_format_from_pdf(pdf, return_format_list=True)

                cur_trm.documents[doc_name].extend([file_size, format_list, status])
                doc_dict[doc_num].append(format_list)

            bar.print_progress_bar()

        print('Pages info was successfully collected!')
        return total_size

    def __collect_docs_to_be_printed(self, target_dir: str):
        """
        Collects documents for printing in target directory.

        Parameters
        ----------
        target_dir : str
            Path to a folder where documents are to be collected for printing.

        Returns
        -------
        None
        """
        item_names_list = self.db.get_item_names()
        trm_names_list = [item_name for item_name in item_names_list if 'TRM' in item_name]

        for trm_name in trm_names_list:
            trm_id = item_names_list.index(trm_name)
            trm = self.db.get_item(trm_id)

            print('Copying files from {}'.format(trm.name))

            for doc in trm.documents:
                if trm.documents[doc] is not None:
                    if trm.documents[doc][-1] == 'Ок':
                        file_path = os.path.join(trm.path, doc + '.pdf')
                        shutil.copy2(file_path, target_dir)

        print('Copying files was successfully completed!')

    def __write_docs_info_to_be_printed(self, doc_dict: dict, target_dir: str):
        """
        Writes document information (document number, number of pages of each format, transmittal, document status)
        to an excel-file.

        Parameters
        ----------
        doc_dict : dict
            Documents dictionary with parsed information.
        target_dir : str
            Path to a folder where documents are to be collected for printing.

        Returns
        -------
        None
        """
        format_dict = self.__cfg[3]

        print('Writing files info into sheet count file...')
        template_path = os.path.join(self.__cfg[0]['templates_dir'], 'sheet_count_template.xlsx')
        wb = load_workbook(template_path)
        xlsheet = wb['Sheet1']

        pages_total_dict = dict.fromkeys(format_dict.keys())
        for key in pages_total_dict.keys():
            pages_total_dict[key] = 0

        pages_total_dict['Other'] = 0
        i_shifted = 0

        total = len(doc_dict)
        bar = PrintProgressBar(start=0, total=total, prefix='Progress:', suffix='Complete', length=50)

        for i, doc_num in enumerate(doc_dict):
            i_shifted = i + 3
            xlsheet.cell(row=i_shifted, column=1).value = doc_num
            xlsheet.cell(row=i_shifted, column=8).value = doc_dict[doc_num][0]
            xlsheet.cell(row=i_shifted, column=9).value = doc_dict[doc_num][2]

            if len(doc_dict[doc_num]) == 4:
                format_list = doc_dict[doc_num][3]
                pages_count_dict = {k: format_list.count(k) for k in pages_total_dict.keys()}

                for j, k in enumerate(pages_total_dict.keys()):
                    pages_total_dict[k] += pages_count_dict[k]
                    xlsheet.cell(row=i_shifted, column=j + 2).value = pages_count_dict[k]

            bar.print_progress_bar()

        xlsheet.cell(row=i_shifted + 1, column=1).value = 'Итого'
        for j, k in enumerate(pages_total_dict.keys()):
            xlsheet.cell(row=i_shifted + 1, column=j + 2).value = pages_total_dict[k]

        target_path = os.path.join(target_dir, 'sheet_count.xlsx')
        wb.save(target_path)
        print('Writing files info was successfully completed!')

    def prepare_docs_for_printing(self, cur_vdr: Document):
        """
        Performs all preparations needed for printing documents from received transmittals.

        Parameters
        ----------
        cur_vdr : Document
            Current VDR.

        Returns
        -------
        None
        """
        target_dir = self.__print_dir
        doc_dict = self.__parse_all_docs_info_from_vdr(cur_vdr)

        total_size = self.__parse_docs_in_received_trms(doc_dict)
        target_free = shutil.disk_usage(target_dir).free

        if target_free > total_size + 100*1024**2:
            self.__collect_docs_to_be_printed(target_dir)
        else:
            print('ERROR: disk space is not enough for copying files!', file=sys.stderr)
        self.__write_docs_info_to_be_printed(doc_dict, target_dir)
