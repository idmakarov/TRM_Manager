"""

This script performs...

"""

import os
import fnmatch
import datetime
import pandas as pd

from openpyxl import load_workbook

# Список ревизий поставщика для пересмотра документа
ifr_list=['A1', 'B1', 'C1', 'D1', 'E1', 'F1', 'G1', 'H1']
# Номера столюцов с датами отсылки TRM с целью IFR
ifr_cols = [66, 103, 112, 121, 130, 139, 148, 157]
# Список ревизий поставщика для использования документа
ifu_list=['00', '01', '02', '03', '04', '05']
# Номера столюцов с датами отсылки TRM с целью IFU
ifu_cols = [166, 175, 184, 193, 202, 211]

# Пропишем пути
#remote_dir = r'berlin\se\УКП\_Общая\Зубков\__AGPP\0002. PU\Отправленные трансмиттелы\0055-P2-GA1-CPC-TRM-00572_'
templates_dir = 'templates'
temporary_dir = 'prc'
remote_dir = 'remote_directory'
subfolder = 'remote_directory/0055-P2-GA1-CPC-TRM-00572_'

# Сформируем список документов на основе pdf-файлов в трансмиттеле
file_list = []
mask = '*.pdf'
for item in os.listdir(subfolder):
    if fnmatch.fnmatch(item, mask):
#        Создадим шаблон, по которому будем искать нужный VDR
        phase_mask = item.split('.')[0] + '.' + item.split('.')[1] +\
        '*' + '.xls*'
#        Отбросим ненужную часть названия документа '.pdf'
        doc_name = item.split('.pdf')[0]
        file_list.append(doc_name)
# Создадим словарь, в который дальше будем записывать данные по каждому документу
file_dict = dict.fromkeys(file_list)

# Найдём путь к нужному VDR
vdr_src = ''
for item in os.listdir(remote_dir):
    if fnmatch.fnmatch(item, phase_mask):
        vdr_src = os.path.join(remote_dir, item)
# Считаем файл VDR
workbook = pd.ExcelFile(vdr_src)
# Распарсим лист в датафрейм
df = workbook.parse(sheet_name='VDR')
# Укажем файл и драйвер записи
vdr_tmp = os.path.join(temporary_dir,'vdr.xlsx')
writer = pd.ExcelWriter(vdr_tmp, engine='xlsxwriter')
# Запишем датафрейм в эксель-файл, на лист VDR
df.to_excel(writer, 'VDR')
# Сохраним файл
writer.save()

# Загрузим данный VDR, при этом считываем только значения в ячейках
wb = load_workbook(vdr_tmp, data_only=True)
sheet = wb['VDR']

def get_vdr_ind(xlsheet, doc_name):
    for i in range(13, xlsheet.max_row+1):
        cur_cell_value = xlsheet.cell(row=i, column=43).value
        if not cur_cell_value is None and cur_cell_value == doc_name:
            return i

# Функция получения имени трансмиттела из абсолютного пути к нему
# Should to be edited: this function must perform simple filtration of TRM name.
# Upper parser of all TRM folders must be implemented.
def get_trm_name(trm_path: str):
    ap = os.path.abspath(trm_path)
    trm_name = os.path.split(ap)[-1]
    if '_' in trm_name:
        trm_name = trm_name.split('_')[0]
    elif ' ' in trm_name:
        trm_name = trm_name.split(' ')[0]
    return trm_name

# Функция для формирования списка дат
def get_date_list(issue_list, issue_cols):
    date_list_ = []
    ind = issue_list.index(revision) + 1
    for date_col in issue_cols[:ind]:
        rev_date = sheet.cell(row=vdr_ind, column=date_col).value
        if not rev_date is None:
            try:
                date_list_.append(rev_date.strftime('%d.%m.%Y'))
            except AttributeError:
                if not (rev_date == '' or rev_date == '-' or rev_date is None):
                    print('Ошибка! Значение ячейки {} в VDR не является датой. \
                          Убедитесь, что Вы ввели дату и поменяйте формат на "Дата".')
    return date_list_

# Получим имя трансмиттела
trm_name = get_trm_name(subfolder)
# Заполним словарь необходимой информацией по каждому документу
for doc in file_list:
    info_list = []
    vdr_ind = get_vdr_ind(sheet, doc)
# Цель выпуска документа
    info_list.append(sheet.cell(row=vdr_ind, column=56).value)
# Класс документа
    info_list.append(sheet.cell(row=vdr_ind, column=23).value)
# Наименование документа (рус.)
    info_list.append(sheet.cell(row=vdr_ind,column=51).value)
# Наименование документа (англ.)
    info_list.append(sheet.cell(row=vdr_ind,column=50).value)
# Ревизия поставщика
    info_list.append(sheet.cell(row=vdr_ind,column=55).value)
# Номер документа
    info_list.append(sheet.cell(row=vdr_ind,column=42).value)
# Соберём список дат отправленных трансмиттелов до текущей ревизии.
    revision = sheet.cell(row=vdr_ind, column=55).value
# Если номер ревизии не букво-цифровой, то искать в столбцах 00, 01,..
    if str(revision) in ifu_list:
        date_list = get_date_list(ifu_list, ifu_cols)
# Иначе искать в столбцах А1, В1,..
    else:
        date_list = get_date_list(ifr_list, ifr_cols)
    if len(date_list) > 0:
        info_list.append(date_list)
    else:
        info_list.append(None)
# Записываем информацию по текущему документу в созданный словарь
    file_dict[doc] = info_list
    
print(file_dict)
